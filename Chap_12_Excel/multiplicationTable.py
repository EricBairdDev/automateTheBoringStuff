import openpyxl

def create_multiplication_table(n):
    # Create a new workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Add the headers for the table
    for i in range(1, n + 1):
        cell = worksheet.cell(row=1, column= i + 1)
        cell.value = i
        cell.font = cell.font.copy(bold=True)

    for i in range(1, n + 1):
        cell = worksheet.cell(row = i + 1, column=1)
        cell.value = i
        cell.font = cell.font.copy(bold=True)

    # Fill in the rest of the table with the products of each row and column
    for row in range(2, n + 2):
        for col in range(2, n + 2):
            worksheet.cell(row=row, column=col).value = (row - 1) * (col - 1)

    # Save the workbook
    workbook.save("multiplication_table.xlsx")

# Get the number N from the command line
import sys
n = int(sys.argv[1])

create_multiplication_table(n)
